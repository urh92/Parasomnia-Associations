# Import necessary libraries
import os
import pandas as pd
import math
import numpy as np
import matplotlib.pyplot as plt
from statsmodels.miscmodels.ordinal_model import OrderedModel
from scipy.stats import ttest_ind, chi2_contingency
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Suppress SettingWithCopyWarning
pd.options.mode.chained_assignment = None

def flatten(l):
    """
    Flatten a list of lists into a single list.

    Args:
        l (list of list): List of lists to be flattened.

    Returns:
        list: A single flattened list.
    """
    return [item for sublist in l for item in sublist]

def fill_nan(row):
    """
    Fill a single NaN value in a row based on other values summing to 100.

    Args:
        row (pd.Series): A pandas Series representing a row from a DataFrame.

    Returns:
        pd.Series: The row with NaN value filled.
    """
    if row.isna().sum() == 1:
        nan_col = row.index[row.isna()][0]
        row[nan_col] = 100 - row.dropna().sum()
    return row

def filter_df_for_ahi_features(df):
    """
    Filter DataFrame for rows meeting specified AHI feature conditions.

    Args:
        df (pd.DataFrame): The DataFrame to be filtered.

    Returns:
        pd.DataFrame: The filtered DataFrame.
    """
    return df.query('Duration_REM >= 10 and Duration_NREM >= 10 and TST >= 60')

def calculate_delta(df, col_rem, col_nrem, col_new):
    """
    Calculate the log delta between REM and NREM indices.

    Args:
        df (pd.DataFrame): The DataFrame containing the data.
        col_rem (str): The column name for REM indices.
        col_nrem (str): The column name for NREM indices.
        col_new (str): The name for the new column to store the log delta values.

    Returns:
        pd.DataFrame: The DataFrame with the new column added.
    """
    df = df.dropna(subset=[col_rem, col_nrem])
    df = df[df[col_new.replace('_Delta', '')] != 0]
    df = df[(df[col_rem] >= 5) | (df[col_nrem] >= 5)]
    df[col_new] = np.log(df[col_rem] + 1) - np.log(df[col_nrem] + 1)
    return df

def calculate_ratio(df, col_numerator, col_denominator, col_new):
    """
    Calculate ratio between two columns.

    Args:
        df (pd.DataFrame): The DataFrame containing the data.
        col_numerator (str): The numerator column name.
        col_denominator (str): The denominator column name.
        col_new (str): The name for the new column to store the ratio.

    Returns:
        pd.DataFrame: The DataFrame with the new column added.
    """
    df = df.dropna(subset=[col_numerator, col_denominator])
    df = df[df[col_denominator] != 0]
    df[col_new] = df[col_numerator] / df[col_denominator]
    return df

def calc_ahi_features(df, col):
    """
    Apply AHI feature calculations based on the specified column.

    Args:
        df (pd.DataFrame): The DataFrame to process.
        col (str): The column name based on which the calculation will be applied.

    Returns:
        pd.DataFrame: The DataFrame with applied calculations.
    """
    df = filter_df_for_ahi_features(df)
    feature_calculations = {
        'AHI_Delta': lambda df: calculate_delta(df, 'Indices_AHI_REM', 'Indices_AHI_NREM', 'AHI_Delta'),
        'AI_Delta': lambda df: calculate_delta(df, 'Indices_AI_REM', 'Indices_AI_NREM', 'AI_Delta'),
        'Percentage_REM_Ratio': lambda df: calculate_ratio(df, 'Percentage_REM', 'Percentage_NREM', 'Percentage_REM_Ratio')
    }
    if col in feature_calculations:
        df = feature_calculations[col](df)
    return df

def replace_nan(df, col, col2):
    """
    Replace NaN values in one column with values from another column if they exist.

    Args:
        df (pd.DataFrame): The DataFrame containing the data.
        col (str): The primary column name to check for NaN values.
        col2 (str): The secondary column name from which to take values if NaN is found in the primary column.

    Returns:
        pd.DataFrame: The DataFrame with NaN values replaced.
    """
    df[col] = np.where(df[col].isna() & df[col2].notna(), df[col2], df[col])
    return df

def calc_time(df, col, start_time=0):
    """
    Calculate time relative to a start time and convert to a circular time representation.

    Args:
        df (pd.DataFrame): The DataFrame containing the time data.
        col (str): The column name containing time data.
        start_time (int, optional): The start time to calculate relative time from. Defaults to 0.

    Returns:
        pd.DataFrame: The DataFrame with the time column converted to circular time representation.
    """
    df[col] = pd.to_datetime(df[col])
    df['time_of_day'] = (df[col] - df[col].dt.normalize() - pd.Timedelta(hours=start_time)).dt.total_seconds()
    df['cos_time'] = np.cos(2 * np.pi * df['time_of_day'] / (24 * 60 * 60))
    df['sin_time'] = np.sin(2 * np.pi * df['time_of_day'] / (24 * 60 * 60))
    df[col] = np.arctan2(df['sin_time'], df['cos_time'])
    df = df.drop(columns=['time_of_day', 'cos_time', 'sin_time'])
    return df

def convert_sleep_hours(s):
    """
    Converts various textual representations of sleep durations into a single float value indicating total hours.

    Args:
        s (str): The string representation of sleep hours.

    Returns:
        float or np.nan: The converted sleep hours as a float value, or np.nan for unhandled formats or missing data.
    """
    # Check for missing data or specific non-numeric strings that cannot be directly converted.
    if pd.isna(s) or s in ['?', 'varies', '10 - 15 minutes', '4-5 minutes', '15 MINUTE', '-']:
        return np.nan  # Return np.nan for these cases to indicate missing or unconvertible data.
    
    # Directly convert digit-only strings to float values.
    if s.isdigit():
        return float(s)  # Convert strings that are purely digits into floats.
    
    # Define patterns for sleep duration representations that need conversion.
    # Each pattern maps to a specific numeric value.
    hour_patterns = {
        r'4\s*-\s*41/2\s*(hrs|hours)?': 4.5,  # Pattern for ranges like "4-41/2 hours" to be converted to 4.5 hours.
        r'5\s*-\s*hr': 5,  # Pattern for strings like "5-hr" indicating 5 hours.
        r'51/2`?|5\s*-\s*6\??|5\s*hours,\s*occasionally\s*7\s*-\s*8|5\s*-\s*6\s*Times': 5.5,  # Patterns for 5.5 hours.
        r'5\s*1/2\s*-\s*6\s*hours|5\s*-\s*6\s*1/2\s*hrs\.?|4\s*-\s*8\s*interrupted|\s*-\s*6\s*hr': 6,  # Patterns for 6 hours.
        r'6\s*-\s*61/2\s*hrs\.?|6\s*-\s*7X|6\s*--\s*7\s*HOURS|6\s*-\s*7\s*times|6\s*-\s*6\s*1/2|6\s*-\s*7\s*now|6\s*--\s*7\s*hours': 6.5,  # Patterns for 6.5 hours.
    }
    
    # Check the input string against each pattern and return the corresponding value if a match is found.
    for pattern, value in hour_patterns.items():
        if re.search(pattern, s, re.IGNORECASE):
            return value  # Return the value associated with the matched pattern.
    
    # For general cases not covered by specific patterns above, use a regular expression to find ranges.
    # This handles strings representing a range of hours, like "7 plus", "7-8 total", "8-1/2hrs."
    range_match = re.search(r'(\d+)[\s-]*(\d+)?(?:\s*/\s*2)?(?:\s*hours?|\s*hrs\.?|h)?', s, re.IGNORECASE)
    if range_match:
        start, end = range_match.groups()  # Extract the start and end values of the range, if present.
        if end:
            return (float(start) + float(end)) / 2  # Calculate the average of the range if an end value is present.
        return float(start)  # Return the start value as a float if no end value is specified.
    
    # Return np.nan for any formats not handled by the above logic.
    return np.nan
    
def format_to_three_sig_figs(value):
    """
    Format a numeric value to three significant figures, handling special cases.
    
    Args:
    value (float): Numeric value to be formatted.
    
    Returns:
    str: Formatted string with three significant figures or a special notation.
    """
    if np.isnan(value):
        return value
    if value == 0.0 or value < 1.00e-300:
        return "<1e-300"
    elif value >= 1:
        return "> 0.999"
    elif value <= 1 and value >= 0.1:
        return "{:.3f}".format(value)
    elif value <= 0.1 and value >= 0.01:
        return "{:.4f}".format(value)
    elif value <= 0.01 and value >= 0.001:
        return "{:.5f}".format(value)
    else:
        return "{:.2e}".format(value)

def display_values(value):
    """
    Display numeric values with appropriate precision based on their magnitude.
    
    Args:
    value (float): Numeric value to be displayed.
    
    Returns:
    str: Formatted string representation of the numeric value.
    """
    if value < 1:
        return "{:.3f}".format(value)
    elif value >= 1 and value < 10:
        return "{:.2f}".format(value)
    elif value >= 10 and value < 100:
        return "{:.1f}".format(value)
    elif value >= 100 and value < 1000:
        return "{:.0f}".format(value)

def custom_format(value):
    """
    Custom format for numeric values based on their absolute value.
    
    Args:
    value (float): Numeric value to be formatted.
    
    Returns:
    str: Formatted string representation of the numeric value.
    """
    abs_val = abs(value)
    if abs_val >= 100:
        return str(int(value))
    elif abs_val >= 10:
        return "{:.1f}".format(value)
    else:
        return "{:.2f}".format(value)

def convert_pvalue_to_float(pval_str):
    """
    Convert a p-value from string format to float, especially handling scientific notation.
    
    Args:
    pval_str (str): P-value string, possibly in scientific notation.
    
    Returns:
    float: Numeric representation of the p-value, or NaN if conversion fails.
    """
    if "e" in pval_str:
        base, exponent = pval_str.split('e')
        try:
            return float(base) * (10 ** int(exponent))
        except ValueError:
            return np.nan
    else:
        try:
            return float(pval_str)
        except ValueError:
            return np.nan

def calculate_significance_and_avg_p(metric_rows):
    """
    Calculate the total number of significant p-values and their average from a DataFrame.
    
    Args:
    metric_rows (pd.DataFrame): DataFrame containing p-value columns.
    
    Returns:
    tuple: Total number of significant p-values, average p-value.
    """
    p_value_cols = [col for col in metric_rows.columns if 'P-value' in col]
    total_significant = sum(metric_rows[p_value_cols].applymap(lambda x: x < 0.05).sum(axis=1))
    avg_p_value = metric_rows[p_value_cols].mean().mean()
    return total_significant, avg_p_value

def format_with_commas(num):
    """
    Format a number with commas as thousands separators.
    
    Args:
    num (int, float): The number to format.
    
    Returns:
    str: The formatted number with commas.
    """
    return "{:,}".format(num)

def extract_pval_components(p_val):
    """
    Extract the base number and exponent from a p-value in scientific notation.
    
    Args:
    p_val (str): P-value string in scientific notation.
    
    Returns:
    tuple: Base number and exponent as floats, or original value and large negative value if not in scientific notation.
    """
    if 'e-' in str(p_val):
        number, exponent = str(p_val).split('e-')
        return float(number), int(exponent)
    return p_val, -100000

def is_significant(value_tuple):
    """
    Determine if a p-value is considered significant based on its exponent in scientific notation.
    
    Args:
    value_tuple (tuple): Tuple containing the base number and exponent of the p-value.
    
    Returns:
    bool: True if the p-value is considered significant, False otherwise.
    """
    number, exponent = value_tuple
    if exponent in [1, 2, 3]:
        return False
    elif exponent == 4:
        return number <= 9
    elif exponent >= 5:
        return True
    return False

def replace_large_exponent(val):
    """
    Replace large exponent values in strings with a more readable format.
    
    Args:
    val (str): The string possibly containing a large exponent.
    
    Returns:
    str: A readable version of the string with large exponents replaced.
    """
    return '<1.00e-300' if isinstance(val, str) and val.split('e')[-1] == '-300' else val

def merge_rows(f_name, n_row, dataframe):
    """
    Merge specific rows in an Excel file for improved readability.
    
    Args:
    f_name (str): Filename for the Excel file.
    n_row (int): Number of rows to merge.
    dataframe (pd.DataFrame): DataFrame to write to the Excel file.
    """
    path = 'C:/Users/UmaerHANIF/OneDrive - BioSerenity/Documents/Parasomnia Paper/Excel'
    filename = os.path.join(path, f_name)
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        dataframe.to_excel(writer, sheet_name='Sheet1', index=False)
    wb = load_workbook(filename)
    ws = wb.active
    last_row = ws.max_row
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
    for i in range(2, last_row, n_row):
        end_row = min(i + n_row - 1, last_row)
        ws.merge_cells(start_row=i, start_column=1, end_row=end_row, end_column=1)
    wb.save(filename)
    
def insomnia_score(df):
    """
    Calculate and assign insomnia scores to a DataFrame based on specific criteria.
    
    Args:
    df (pd.DataFrame): DataFrame containing insomnia-related columns.
    
    Returns:
    pd.DataFrame: DataFrame with calculated insomnia scores.
    """
    is_dict = {'Not at all / None': 0, 'Slight / Few times': 1, 'Moderate / Sometimes': 2, 'Often': 3, 'Severe / Always': 4}
    df[is_cols] = df[is_cols].replace(is_dict)
    df[is_cols2] = df[is_cols2].replace(is_dict)
    df['Score'] = df[is_cols].sum(1)
    df['Score2'] = df[is_cols2].sum(1)
    df['Insomnia'] = df.apply(lambda row: 1 if (row['Score'] >= 10) and (row['Score2'] >= 9) else 0 if row['Score'] <= 2 else np.nan, axis=1)
    return df
    
def compute_score(df, cols, label):
    """
    Compute a composite score for a set of columns and assign it to a new column in the dataframe.

    Args:
        df (pd.DataFrame): The dataframe containing the columns to be scored.
        cols (list): List of column names to be included in the score computation.
        label (str): The name of the new column where the computed score will be stored.

    Returns:
        pd.DataFrame: The dataframe with the new score column added.
    """
    # Define a dictionary to map response options to numeric scores.
    score_dict = {'Not at all / None': 0, 'Slight / Few times': 1, 'Moderate / Sometimes': 2, 'Often': 3, 'Severe / Always': 4}
    df2 = df.copy()  # Create a copy of the dataframe to avoid modifying the original data.
    
    # Replace the response options in the specified columns with their corresponding numeric scores.
    for col in cols:
        if 'Not at all / None' in df2[col].unique().tolist():
            df2[col] = df2[col].replace(score_dict)
        else:
            # If the column does not include the 'Not at all / None' option, replace a specific value with 3.
            df2[col] = df2[col].replace(1, 3)
    
    # Sum the scores across the specified columns and assign the result to the new column.
    df[label] = df2[cols].sum(1)
    return df

def clean_data(dataframe, column, dichotimize=False):
    """
    Clean and preprocess data in a specified column of the dataframe based on predefined criteria.

    Args:
        dataframe (pd.DataFrame): The dataframe to be cleaned.
        column (str): The name of the column to be cleaned.
        dichotimize (bool, optional): Indicates whether to apply dichotomization to the column. Defaults to False.

    Returns:
        pd.DataFrame: The cleaned dataframe.
    """
    # Define criteria for cleaning specific columns.
    column_criteria = {
        'Gender': lambda df: df.replace({'Male': 1, 'Female': 0}),
        'BMI': lambda df: df[df[column] <= 80],
        'TIB': lambda df: df[df[column] <= 600],
        'WASO': lambda df: df[df[column] <= 500],
        'Latency_REM': lambda df: df[df[column] <= 300].abs(),
        'SEN': lambda df: df[df[column] <= 6000],
        'AHI_Delta': lambda df: df,  # No specific cleaning, just return the dataframe.
        'Percentage_REM_Ratio': lambda df: df  # No specific cleaning, just return the dataframe.
    }

    # Define groups of columns with similar cleaning operations based on bounds.
    bounds_checks = {
        ('LowSat', 'Minimum_SpO2', 'Baseline_SpO2', 'Mean_SpO2'): (60, 100),
        ('SAO2_Per', 'TST_90', 'TST_80', 'TST_70'): (None, 100),
        ('Indices_AHI_Total', 'Indices_AHI_NREM', 'Indices_AHI_REM', 'RDI', 'RDI_NREM', 'RDI_REM', 'AI_NREM', 'AI_REM', 
         'Indices_AI_NREM', 'Indices_AI_REM', 'AI_Total', 'AI_Total2', 'AI_Spontaneous', 'AI_Respiratory', 'AI_PLMS', 
         'AI_Micro', 'AI_Awakenings', 'PLMS'): (None, 200),
        ('Durations_OA_Average_Duration', 'Durations_HI_Average_Duration', 'OA_HYP_Duration'): (10, None),
        'Mean_SpO2_90': (60, 90),
        ('PH3_Sleep_Hours', 'ESS'): (None, 24)
    }

    # Apply the predefined criteria to clean the specified column.
    if column in column_criteria:
        dataframe = column_criteria[column](dataframe)
    else:
        # For columns that require checking against bounds, apply the appropriate filtering.
        for cols, (min_val, max_val) in bounds_checks.items():
            if column in cols:
                if min_val is not None and max_val is not None:
                    dataframe = dataframe[(dataframe[column] >= min_val) & (dataframe[column] <= max_val)]
                elif min_val is not None:
                    dataframe = dataframe[dataframe[column] >= min_val]
                elif max_val is not None:
                    dataframe = dataframe[dataframe[column] <= max_val]
                break
        else:  # Handle unique cases or apply default numeric conversion
            if 'Not at all / None' in dataframe[column].unique().tolist():
                replace_dict = others_dict if dichotimize else anxiety_dict
                dataframe[column] = dataframe[column].replace(replace_dict)
            elif column in ['PostSleep_Feeling_Refreshing', 'PostSleep_Awake_Feeling', 'PostSleep_Last_Night_Sleep_Quality']:
                replace_dict = {
                    'PostSleep_Feeling_Refreshing': refreshed_dict,
                    'PostSleep_Awake_Feeling': awake_feeling_dict,
                    'PostSleep_Last_Night_Sleep_Quality': sleep_quality_dict
                }
                dataframe[column] = dataframe[column].replace(replace_dict[column])
            else:
                dataframe[column] = pd.to_numeric(dataframe[column], errors='coerce').abs()

    # Apply log transformation for specific columns
    if column in log_cols:
        dataframe[column] = np.log(dataframe[column] + 1)
    return dataframe
